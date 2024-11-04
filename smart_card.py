import openpyxl
from BX24 import Bitrix24
from bitrix_candidate import parse_webhook, BITRIX_WEBHOOK_URL


def  get_data_from_xls():
    workbook = openpyxl.load_workbook('./database.xlsx')
    worksheet = workbook.active

    dictionary = {}
    
    for row in range(1, worksheet.max_row + 1):
      key = worksheet.cell(row, 1).value
      value = worksheet.cell(row, 2).value
      dictionary[key] = value
    return dictionary

candidate_id=1

def create_smart_card(candidate_id):
    domain, webhook_key = parse_webhook(BITRIX_WEBHOOK_URL)
    bx24 = Bitrix24(domain, webhook_key, candidate_id)
    body = {
        "entityTypeId": 1,
        'fields': get_data_from_xls(),
        'params': {
            "REGISTER_SONET_EVENT": "Y",
        }
    } 
    bx24.call('crm.item.add', body)
    
create_smart_card(candidate_id)

